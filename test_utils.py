import json
import os
import socket
import subprocess
import tempfile
import time
import uuid

import numpy as np
import pandas as pd
import pytest
import pydicom
import requests
from pydicom.dataset import FileDataset, FileMetaDataset
from pydicom.uid import generate_uid

from utils import csv_string_to_xlsx, Spreadsheet, generate_queries_and_filter, save_failed_queries_csv, find_studies_from_pacs_list, get_studies_from_study_pacs_map, PacsConfiguration


def _create_test_dicom(accession, patient_id, patient_name, modality, slice_thickness):
    ds = Fixtures.create_minimal_dicom(
        patient_id=patient_id,
        patient_name=patient_name,
        accession=accession,
        study_date="20250101",
        modality=modality,
        SliceThickness=slice_thickness
    )
    ds.InstitutionName = "Test Hospital"
    ds.ReferringPhysicianName = "Dr. Referring"
    ds.Manufacturer = "TestManufacturer"
    ds.ManufacturerModelName = "TestModel"
    ds.SeriesNumber = 1
    ds.SamplesPerPixel = 1
    ds.PhotometricInterpretation = "MONOCHROME2"
    ds.Rows = 64
    ds.Columns = 64
    ds.BitsAllocated = 16
    ds.BitsStored = 16
    ds.HighBit = 15
    ds.PixelRepresentation = 0
    ds.PixelData = np.random.randint(0, 1000, (64, 64), dtype=np.uint16).tobytes()
    return ds


def _upload_dicom_to_orthanc(ds, orthanc):
    temp_file = tempfile.mktemp(suffix=".dcm")
    ds.save_as(temp_file)
    orthanc.upload_dicom(temp_file)
    os.remove(temp_file)
    # Give time for the DICOM to be uploaded to Orthanc
    time.sleep(2)


def test_csv_string_to_xlsx_basic(tmp_path):
    csv_string = """Name,Age,City
John,30,NYC
Jane,25,LA
Bob,35,Chicago"""
    
    output_file = tmp_path / "output.xlsx"
    csv_string_to_xlsx(csv_string, str(output_file))
    
    assert output_file.exists()
    
    df = pd.read_excel(output_file)
    assert len(df) == 3
    assert list(df.columns) == ["Name", "Age", "City"]
    assert df.loc[0, "Name"] == "John"
    assert df.loc[1, "Name"] == "Jane"
    assert df.loc[2, "Name"] == "Bob"
    assert str(df.loc[0, "Age"]) == "30"
    assert str(df.loc[1, "Age"]) == "25"


def test_csv_string_to_xlsx_ctp_format_with_parens(tmp_path):
    csv_string = """AccessionNumber,PatientID,Status
=("ACC123"),=("MRN456"),=("Complete")
=("ACC124"),=("MRN457"),=("Pending")"""
    
    output_file = tmp_path / "output.xlsx"
    csv_string_to_xlsx(csv_string, str(output_file))
    
    df = pd.read_excel(output_file)
    assert len(df) == 2
    assert df.loc[0, "AccessionNumber"] == "ACC123"
    assert df.loc[0, "PatientID"] == "MRN456"
    assert df.loc[0, "Status"] == "Complete"
    assert df.loc[1, "AccessionNumber"] == "ACC124"


def test_csv_string_to_xlsx_ctp_format_with_quotes(tmp_path):
    csv_string = """AccessionNumber,PatientID
="ACC123",="MRN456"
="ACC124",="MRN457" """
    
    output_file = tmp_path / "output.xlsx"
    csv_string_to_xlsx(csv_string, str(output_file))
    
    df = pd.read_excel(output_file)
    assert len(df) == 2
    assert df.loc[0, "AccessionNumber"] == "ACC123"
    assert df.loc[0, "PatientID"] == "MRN456"
    assert df.loc[1, "AccessionNumber"] == "ACC124"
    assert df.loc[1, "PatientID"] == "MRN457"


def test_csv_string_to_xlsx_date_yyyymmdd(tmp_path):
    csv_string = """Name,StudyDate,Value
John,20250101,100
Jane,20250102,200"""
    
    output_file = tmp_path / "output.xlsx"
    csv_string_to_xlsx(csv_string, str(output_file))
    
    df = pd.read_excel(output_file)
    assert len(df) == 2
    assert isinstance(df.loc[0, "StudyDate"], pd.Timestamp)
    assert df.loc[0, "StudyDate"].year == 2025
    assert df.loc[0, "StudyDate"].month == 1
    assert df.loc[0, "StudyDate"].day == 1


def test_csv_string_to_xlsx_date_yyyy_mm_dd(tmp_path):
    csv_string = """Name,StudyDate,Value
John,2025-01-01,100
Jane,2025-01-02,200"""
    
    output_file = tmp_path / "output.xlsx"
    csv_string_to_xlsx(csv_string, str(output_file))
    
    df = pd.read_excel(output_file)
    assert isinstance(df.loc[0, "StudyDate"], pd.Timestamp)
    assert df.loc[0, "StudyDate"].year == 2025


def test_csv_string_to_xlsx_date_mm_dd_yyyy(tmp_path):
    csv_string = """Name,BirthDate,Value
John,01/15/2025,100
Jane,02/20/2025,200"""
    
    output_file = tmp_path / "output.xlsx"
    csv_string_to_xlsx(csv_string, str(output_file))
    
    df = pd.read_excel(output_file)
    assert isinstance(df.loc[0, "BirthDate"], pd.Timestamp)
    assert df.loc[0, "BirthDate"].year == 2025
    assert df.loc[0, "BirthDate"].month == 1
    assert df.loc[0, "BirthDate"].day == 15


def test_csv_string_to_xlsx_date_case_insensitive(tmp_path):
    csv_string = """Name,STUDYDATE,SeriesDate,Value
John,20250101,20250102,100"""
    
    output_file = tmp_path / "output.xlsx"
    csv_string_to_xlsx(csv_string, str(output_file))
    
    df = pd.read_excel(output_file)
    assert isinstance(df.loc[0, "STUDYDATE"], pd.Timestamp)
    assert isinstance(df.loc[0, "SeriesDate"], pd.Timestamp)


def test_csv_string_to_xlsx_mixed_formats(tmp_path):
    from openpyxl import load_workbook
    
    csv_string = """AccessionNumber,PatientID,StudyDate,Value
=("ACC123"),="MRN456",20250101,="100"
=("ACC124"),="MRN457",20250102,="200" """
    
    output_file = tmp_path / "output.xlsx"
    csv_string_to_xlsx(csv_string, str(output_file))
    
    df = pd.read_excel(output_file)
    assert df.loc[0, "AccessionNumber"] == "ACC123"
    assert df.loc[0, "PatientID"] == "MRN456"
    assert isinstance(df.loc[0, "StudyDate"], pd.Timestamp)
    
    wb = load_workbook(output_file)
    ws = wb.active
    assert ws.cell(row=2, column=1).number_format == '@', "AccessionNumber should be text format"
    assert ws.cell(row=2, column=2).number_format == '@', "PatientID should be text format"
    assert ws.cell(row=2, column=4).number_format == '@', "Value should be text format"
    assert ws.cell(row=2, column=1).value == "ACC123"
    assert ws.cell(row=2, column=4).value == "100"


def test_csv_string_to_xlsx_empty_string(tmp_path):
    csv_string = ""
    
    output_file = tmp_path / "output.xlsx"
    csv_string_to_xlsx(csv_string, str(output_file))
    
    assert output_file.exists()
    df = pd.read_excel(output_file)
    assert len(df) == 0


def test_csv_string_to_xlsx_only_header(tmp_path):
    csv_string = "Name,Age,City"
    
    output_file = tmp_path / "output.xlsx"
    csv_string_to_xlsx(csv_string, str(output_file))
    
    assert output_file.exists()
    df = pd.read_excel(output_file)
    assert len(df) == 0
    assert list(df.columns) == ["Name", "Age", "City"]


def test_csv_string_to_xlsx_all_cells_as_strings(tmp_path):
    from openpyxl import load_workbook
    
    csv_string = """Name,Age,Value
John,30,123
Jane,25,456"""
    
    output_file = tmp_path / "output.xlsx"
    csv_string_to_xlsx(csv_string, str(output_file))
    
    wb = load_workbook(output_file)
    ws = wb.active
    assert ws.cell(row=2, column=1).number_format == '@', "Name should be text format"
    assert ws.cell(row=2, column=2).number_format == '@', "Age should be text format"
    assert ws.cell(row=2, column=3).number_format == '@', "Value should be text format"


def test_csv_string_to_xlsx_invalid_date_stays_string(tmp_path):
    from datetime import datetime
    
    csv_string = """Name,StudyDate,Value
John,NotADate,100
Jane,20250102,200"""
    
    output_file = tmp_path / "output.xlsx"
    csv_string_to_xlsx(csv_string, str(output_file))
    
    df = pd.read_excel(output_file)
    assert isinstance(df.loc[0, "StudyDate"], str)
    assert df.loc[0, "StudyDate"] == "NotADate"
    assert isinstance(df.loc[1, "StudyDate"], (pd.Timestamp, datetime))


class Fixtures:
    
    @staticmethod
    def create_minimal_dicom(patient_id="TEST001", patient_name="DOE^JOHN", 
                            accession="ACC001", study_date="20240101", 
                            modality="CT", **extra_tags):
        file_meta = FileMetaDataset()
        file_meta.MediaStorageSOPClassUID = '1.2.840.10008.5.1.4.1.1.2'
        file_meta.MediaStorageSOPInstanceUID = generate_uid()
        file_meta.TransferSyntaxUID = '1.2.840.10008.1.2.1'
        file_meta.ImplementationClassUID = generate_uid()
        
        ds = FileDataset(None, {}, file_meta=file_meta, preamble=b"\0" * 128)
        
        ds.PatientName = patient_name
        ds.PatientID = patient_id
        ds.AccessionNumber = accession
        ds.StudyDate = study_date
        ds.StudyTime = "120000"
        ds.Modality = modality
        ds.StudyInstanceUID = generate_uid()
        ds.SeriesInstanceUID = generate_uid()
        ds.SOPInstanceUID = file_meta.MediaStorageSOPInstanceUID
        ds.SOPClassUID = file_meta.MediaStorageSOPClassUID
        ds.SeriesNumber = 1
        ds.InstanceNumber = 1
        
        for key, value in extra_tags.items():
            setattr(ds, key, value)
        
        ds.is_little_endian = True
        ds.is_implicit_VR = False
        
        return ds


class OrthancServer:

    def __init__(self, aet="ORTHANC_TEST", http_port=None, dicom_port=None):
        self.aet = aet
        self.http_port = http_port or self._get_free_port()
        self.dicom_port = dicom_port or self._get_free_port()
        self.container = None
        self.network = None
        self.base_url = f"http://localhost:{self.http_port}"
        self.modalities = {}
        self.storage_dir = None
        self.config_dir = None
        
    def _get_free_port(self):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.bind(('', 0))
            s.listen(1)
            port = s.getsockname()[1]
        return port
    
    def add_modality(self, name, aet, ip, port):
        self.modalities[name] = [aet, ip, port]
    
    def start(self):
        container_name = f"orthanc_test_{uuid.uuid4().hex[:8]}"
        network_name = f"orthanc_net_{uuid.uuid4().hex[:8]}"

        # Create Docker network for C-GET support
        subprocess.run(["docker", "network", "create", network_name],
                      check=True, capture_output=True)
        self.network = network_name

        self.config_dir = tempfile.mkdtemp()
        self.storage_dir = tempfile.mkdtemp()
        config_path = os.path.join(self.config_dir, "orthanc.json")

        config = {
            "Name": "OrthancTest",
            "DicomAet": self.aet,
            "DicomPort": 4242,  # Internal port in container
            "HttpPort": 8042,   # Internal port in container
            "RemoteAccessEnabled": True,
            "AuthenticationEnabled": False,
            "DicomAlwaysAllowFind": True,
            "DicomAlwaysAllowGet": True,
            "DicomAlwaysAllowMove": True,
            "DicomCheckCalledAet": False,
            "DicomCheckModalityHost": False,
            "DicomModalities": self.modalities
        }

        with open(config_path, 'w') as f:
            json.dump(config, f, indent=2)

        # Run Orthanc on custom network with port mapping
        subprocess.run([
            "docker", "run", "-d",
            "--name", container_name,
            "--network", network_name,
            "-p", f"{self.http_port}:8042",
            "-p", f"{self.dicom_port}:4242",
            "-v", f"{self.config_dir}:/etc/orthanc:ro",
            "orthancteam/orthanc:latest"
        ], check=True, capture_output=True, text=True)

        self.container = container_name

        # Wait for Orthanc to be ready
        for i in range(30):
            try:
                response = requests.get(f"{self.base_url}/system", timeout=1)
                if response.status_code == 200:
                    break
            except:
                pass
            time.sleep(1)
        else:
            # Get container logs for debugging
            logs_result = subprocess.run(
                ["docker", "logs", container_name],
                capture_output=True, text=True
            )
            raise RuntimeError(
                f"Orthanc failed to start after 30 seconds. "
                f"Container logs:\n{logs_result.stdout}\n{logs_result.stderr}"
            )
    
    def upload_dicom(self, file_path):
        with open(file_path, 'rb') as f:
            response = requests.post(f"{self.base_url}/instances", files={'file': f})
        return response.status_code == 200
    
    def upload_study(self, patient_id, accession, study_date="20240101", series=None):
        if series is None:
            series = [{'modality': 'CT', 'series_description': 'SERIES1'}]
        
        study_uid = generate_uid()
        
        for series_num, series_info in enumerate(series):
            series_uid = generate_uid()
            modality = series_info.get('modality', 'CT')
            series_desc = series_info.get('series_description', f'SERIES{series_num+1}')
            series_date = series_info.get('series_date')
            
            for instance_num in range(3):
                ds = Fixtures.create_minimal_dicom(
                    patient_id=patient_id,
                    accession=accession,
                    study_date=study_date,
                    modality=modality
                )
                ds.StudyInstanceUID = study_uid
                ds.SeriesInstanceUID = series_uid
                ds.SeriesNumber = series_num + 1
                ds.InstanceNumber = instance_num + 1
                ds.SeriesDescription = series_desc
                if series_date:
                    ds.SeriesDate = series_date
                
                temp_file = tempfile.mktemp(suffix=".dcm")
                ds.save_as(temp_file)
                self.upload_dicom(temp_file)
                os.remove(temp_file)
    
    def get_study_count(self):
        response = requests.get(f"{self.base_url}/studies")
        return len(response.json())
    
    def get_pacs_config(self):
        """Get a PacsConfiguration object for this Orthanc instance."""
        from utils import PacsConfiguration
        return PacsConfiguration(
            host="localhost",
            port=self.dicom_port,
            aet=self.aet
        )

    def stop(self):
        if self.container:
            subprocess.run(["docker", "stop", self.container], capture_output=True, timeout=10)
            subprocess.run(["docker", "rm", self.container], capture_output=True, timeout=10)
            time.sleep(0.5)  # Brief wait to ensure cleanup completes
        if self.network:
            subprocess.run(["docker", "network", "rm", self.network], capture_output=True, timeout=10)


class AzuriteServer:
    """
    Manages an Azurite Docker container for testing Azure Blob Storage functionality.
    Azurite is the official Azure Storage emulator.
    """
    
    def __init__(self, blob_port=None):
        self.blob_port = blob_port or self._get_free_port()
        self.container = None
        self.account_name = "devstoreaccount1"
        self.account_key = "Eby8vdM02xNOcqFlqUwJPLlmEtlCDXJ1OUzFT50uSRZ6IFsuFq2UVErCz4I6tq/K1SZFPTOtr/KBHBeksoGMGw=="
        self.connection_string = (
            f"DefaultEndpointsProtocol=http;"
            f"AccountName={self.account_name};"
            f"AccountKey={self.account_key};"
            f"BlobEndpoint=http://127.0.0.1:{self.blob_port}/{self.account_name};"
        )
        
    def _get_free_port(self):
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.bind(('127.0.0.1', 0))
            s.listen(1)
            port = s.getsockname()[1]
        return port
    
    def start(self):
        """Start the Azurite Docker container"""
        container_name = f"azurite_test_{uuid.uuid4().hex[:8]}"
        
        subprocess.run([
            "docker", "run", "-d",
            "--name", container_name,
            "-p", f"{self.blob_port}:10000",
            "mcr.microsoft.com/azure-storage/azurite:latest",
            "azurite-blob", "--blobHost", "0.0.0.0"
        ], check=True, capture_output=True)
        
        self.container = container_name
        
        # Wait for Azurite to be ready
        time.sleep(2)
        
        # Verify connection
        from azure.storage.blob import BlobServiceClient
        for attempt in range(30):
            try:
                client = BlobServiceClient.from_connection_string(self.connection_string)
                client.get_account_information()
                break
            except Exception:
                if attempt == 29:
                    raise
                time.sleep(1)
    
    def get_sas_url(self, container_name):
        """Generate a SAS URL for a container"""
        from azure.storage.blob import BlobServiceClient, generate_container_sas, ContainerSasPermissions
        from datetime import datetime, timedelta
        
        # Create container if it doesn't exist
        client = BlobServiceClient.from_connection_string(self.connection_string)
        try:
            client.create_container(container_name)
        except Exception:
            pass  # Container might already exist
        
        # Generate SAS token with proper permissions
        sas_token = generate_container_sas(
            account_name=self.account_name,
            container_name=container_name,
            account_key=self.account_key,
            permission=ContainerSasPermissions(read=True, write=True, delete=True, list=True),
            expiry=datetime.utcnow() + timedelta(hours=1)
        )
        
        # Return the container URL (not including container name in path, it's implicit)
        return f"http://127.0.0.1:{self.blob_port}/{self.account_name}/{container_name}?{sas_token}"
    
    def list_blobs(self, container_name):
        """List all blobs in a container"""
        from azure.storage.blob import BlobServiceClient
        
        client = BlobServiceClient.from_connection_string(self.connection_string)
        container_client = client.get_container_client(container_name)
        
        blobs = []
        for blob in container_client.list_blobs():
            blobs.append(blob.name)
        return blobs
    
    def get_blob_content(self, container_name, blob_name):
        """Get the content of a blob"""
        from azure.storage.blob import BlobServiceClient
        
        client = BlobServiceClient.from_connection_string(self.connection_string)
        blob_client = client.get_blob_client(container=container_name, blob=blob_name)
        
        return blob_client.download_blob().readall()
    
    def stop(self):
        """Stop and remove the Azurite Docker container"""
        if self.container:
            subprocess.run(["docker", "stop", self.container], capture_output=True)
            subprocess.run(["docker", "rm", self.container], capture_output=True)


def test_generate_queries_trims_accession_numbers(tmp_path):
    query_file = tmp_path / "query.xlsx"
    query_df = pd.DataFrame({
        "AccessionNumber": ["  ABC001  ", "ABC002", "  ABC003"]
    })
    query_df.to_excel(query_file, index=False)
    
    spreadsheet = Spreadsheet.from_file(str(query_file), acc_col="AccessionNumber")
    query_params_list, expected_values_list, generated_filter = generate_queries_and_filter(spreadsheet)
    
    assert len(query_params_list) == 3
    assert query_params_list[0]["AccessionNumber"] == "*ABC001*"
    assert query_params_list[1]["AccessionNumber"] == "*ABC002*"
    assert query_params_list[2]["AccessionNumber"] == "*ABC003*"
    
    assert len(expected_values_list) == 3
    assert expected_values_list[0] == ("ABC001", 0)
    assert expected_values_list[1] == ("ABC002", 1)
    assert expected_values_list[2] == ("ABC003", 2)


def test_generate_queries_filter_format(tmp_path):
    query_file = tmp_path / "query.xlsx"
    query_df = pd.DataFrame({
        "AccessionNumber": ["ACC001", "ACC002"]
    })
    query_df.to_excel(query_file, index=False)
    
    spreadsheet = Spreadsheet.from_file(str(query_file), acc_col="AccessionNumber")
    query_params_list, expected_values_list, generated_filter = generate_queries_and_filter(spreadsheet)
    
    expected_filter = 'AccessionNumber.contains("ACC001") + AccessionNumber.contains("ACC002")'
    assert generated_filter == expected_filter


def test_generate_queries_mrn_date_no_expected_values(tmp_path):
    query_file = tmp_path / "query.xlsx"
    query_df = pd.DataFrame({
        "AccessionNumber": [None, None],
        "PatientID": ["MRN001", "MRN002"],
        "StudyDate": [pd.Timestamp("2025-01-01"), pd.Timestamp("2025-01-02")]
    })
    query_df.to_excel(query_file, index=False)
    
    spreadsheet = Spreadsheet.from_file(
        str(query_file),
        acc_col="AccessionNumber",
        mrn_col="PatientID",
        date_col="StudyDate"
    )
    query_params_list, expected_values_list, generated_filter = generate_queries_and_filter(spreadsheet)
    
    assert len(expected_values_list) == 0


def test_save_failed_queries_csv_with_accession_numbers(tmp_path):
    query_file = tmp_path / "query.xlsx"
    query_df = pd.DataFrame({
        "AccessionNumber": ["ACC001", "ACC002", "ACC003"]
    })
    query_df.to_excel(query_file, index=False)
    
    spreadsheet = Spreadsheet.from_file(str(query_file), acc_col="AccessionNumber")
    appdata_dir = str(tmp_path / "appdata")
    os.makedirs(appdata_dir, exist_ok=True)
    
    failed_query_indices = [0, 2]
    failure_reasons = {
        0: "Failed to find images",
        2: "Failed to move images after successful query"
    }
    
    save_failed_queries_csv(failed_query_indices, spreadsheet, appdata_dir, failure_reasons)
    
    csv_path = os.path.join(appdata_dir, "failed_queries.csv")
    assert os.path.exists(csv_path)
    
    df = pd.read_csv(csv_path)
    assert len(df) == 2
    assert list(df.columns) == ["Accession Number", "Failure Reason"]
    assert df.loc[0, "Accession Number"] == "ACC001"
    assert df.loc[0, "Failure Reason"] == "Failed to find images"
    assert df.loc[1, "Accession Number"] == "ACC003"
    assert df.loc[1, "Failure Reason"] == "Failed to move images after successful query"


def test_save_failed_queries_csv_with_accession_and_mrn(tmp_path):
    query_file = tmp_path / "query.xlsx"
    query_df = pd.DataFrame({
        "AccessionNumber": ["ACC001", "ACC002", "ACC003"],
        "PatientID": ["MRN001", "MRN002", "MRN003"]
    })
    query_df.to_excel(query_file, index=False)
    
    spreadsheet = Spreadsheet.from_file(
        str(query_file),
        acc_col="AccessionNumber",
        mrn_col="PatientID"
    )
    appdata_dir = str(tmp_path / "appdata")
    os.makedirs(appdata_dir, exist_ok=True)
    
    failed_query_indices = [1]
    failure_reasons = {1: "Failed to find images"}
    
    save_failed_queries_csv(failed_query_indices, spreadsheet, appdata_dir, failure_reasons)
    
    csv_path = os.path.join(appdata_dir, "failed_queries.csv")
    df = pd.read_csv(csv_path)
    assert len(df) == 1
    assert list(df.columns) == ["Accession Number", "MRN", "Failure Reason"]
    assert df.loc[0, "Accession Number"] == "ACC002"
    assert df.loc[0, "MRN"] == "MRN002"
    assert df.loc[0, "Failure Reason"] == "Failed to find images"


def test_save_failed_queries_csv_with_mrn_dates(tmp_path):
    query_file = tmp_path / "query.xlsx"
    query_df = pd.DataFrame({
        "PatientID": ["MRN001", "MRN002", "MRN003"],
        "StudyDate": [
            pd.Timestamp("2025-01-15"),
            pd.Timestamp("2025-02-20"),
            pd.Timestamp("2025-03-10")
        ]
    })
    query_df.to_excel(query_file, index=False)
    
    spreadsheet = Spreadsheet.from_file(
        str(query_file),
        mrn_col="PatientID",
        date_col="StudyDate"
    )
    appdata_dir = str(tmp_path / "appdata")
    os.makedirs(appdata_dir, exist_ok=True)
    
    failed_query_indices = [0, 2]
    failure_reasons = {
        0: "Failed to find images",
        2: "Failed to move images after successful query"
    }
    
    save_failed_queries_csv(failed_query_indices, spreadsheet, appdata_dir, failure_reasons)
    
    csv_path = os.path.join(appdata_dir, "failed_queries.csv")
    df = pd.read_csv(csv_path)
    assert len(df) == 2
    assert list(df.columns) == ["MRN", "Date", "Failure Reason"]
    assert df.loc[0, "MRN"] == "MRN001"
    assert df.loc[0, "Date"] == "2025-01-15"
    assert df.loc[0, "Failure Reason"] == "Failed to find images"
    assert df.loc[1, "MRN"] == "MRN003"
    assert df.loc[1, "Date"] == "2025-03-10"


def test_save_failed_queries_csv_mixed_failure_types(tmp_path):
    query_file = tmp_path / "query.xlsx"
    query_df = pd.DataFrame({
        "AccessionNumber": ["ACC001", "ACC002", "ACC003", "ACC004"]
    })
    query_df.to_excel(query_file, index=False)
    
    spreadsheet = Spreadsheet.from_file(str(query_file), acc_col="AccessionNumber")
    appdata_dir = str(tmp_path / "appdata")
    os.makedirs(appdata_dir, exist_ok=True)
    
    failed_query_indices = [0, 1, 3]
    failure_reasons = {
        0: "Failed to find images",
        1: "Failed to move images after successful query",
        3: "Failed to find images"
    }
    
    save_failed_queries_csv(failed_query_indices, spreadsheet, appdata_dir, failure_reasons)
    
    csv_path = os.path.join(appdata_dir, "failed_queries.csv")
    df = pd.read_csv(csv_path)
    assert len(df) == 3
    assert df.loc[0, "Failure Reason"] == "Failed to find images"
    assert df.loc[1, "Failure Reason"] == "Failed to move images after successful query"
    assert df.loc[2, "Failure Reason"] == "Failed to find images"


def test_save_failed_queries_csv_empty_failures(tmp_path):
    query_file = tmp_path / "query.xlsx"
    query_df = pd.DataFrame({
        "AccessionNumber": ["ACC001", "ACC002"]
    })
    query_df.to_excel(query_file, index=False)
    
    spreadsheet = Spreadsheet.from_file(str(query_file), acc_col="AccessionNumber")
    appdata_dir = str(tmp_path / "appdata")
    os.makedirs(appdata_dir, exist_ok=True)
    
    failed_query_indices = []
    failure_reasons = {}
    
    save_failed_queries_csv(failed_query_indices, spreadsheet, appdata_dir, failure_reasons)
    
    csv_path = os.path.join(appdata_dir, "failed_queries.csv")
    assert os.path.exists(csv_path)
    
    df = pd.read_csv(csv_path)
    assert len(df) == 0
    assert list(df.columns) == ["Accession Number", "Failure Reason"]


def test_find_studies_returns_failure_details(tmp_path):
    orthanc = OrthancServer(aet="ORTHANC")
    try:
        orthanc.start()
        
        orthanc.upload_study(
            patient_id="MRN001",
            accession="ACC001",
            study_date="20250115"
        )
        
        query_file = tmp_path / "query.xlsx"
        query_df = pd.DataFrame({
            "AccessionNumber": ["ACC001", "ACC999"]
        })
        query_df.to_excel(query_file, index=False)
        
        spreadsheet = Spreadsheet.from_file(str(query_file), acc_col="AccessionNumber")
        query_params_list, expected_values_list, _ = generate_queries_and_filter(spreadsheet)
        
        pacs_list = [PacsConfiguration(host="localhost", port=orthanc.dicom_port, aet=orthanc.aet)]
        application_aet = "TEST_AET"
        
        study_pacs_map, failed_query_indices, failure_details = find_studies_from_pacs_list(
            pacs_list, query_params_list, application_aet, expected_values_list
        )
        
        assert len(study_pacs_map) == 1
        assert len(failed_query_indices) == 1
        assert 1 in failed_query_indices
        assert 1 in failure_details
        assert failure_details[1] == "Failed to find images"
        
    finally:
        orthanc.stop()


def test_get_studies_returns_failure_details(tmp_path):
    """Test that get_studies_from_study_pacs_map returns failure details as 3rd value"""
    from unittest.mock import patch, MagicMock

    pacs = PacsConfiguration(host="localhost", port=4242, aet="TEST_PACS")
    study_pacs_map = {
        "study_uid_1": (pacs, 0),
        "study_uid_2": (pacs, 1)
    }

    output_dir = str(tmp_path / "output")

    # Mock get_study to simulate one success and one failure
    with patch('utils.get_study') as mock_get:
        mock_get.side_effect = [
            {"success": True, "num_completed": 5, "num_failed": 0, "num_warning": 0, "message": "Get successful"},
            {"success": False, "num_completed": 0, "num_failed": 0, "num_warning": 0, "message": "Get failed"}
        ]

        successful_gets, failed_query_indices, failure_details = get_studies_from_study_pacs_map(
            study_pacs_map, "TEST_AET", output_dir
        )

        assert successful_gets == 1, "Should have 1 successful get"
        assert len(failed_query_indices) == 1, "Should have 1 failed query index"
        assert 1 in failed_query_indices, "Query index 1 should have failed"
        assert len(failure_details) == 1, "Should have 1 failure detail entry"
        assert 1 in failure_details, "Failure details should contain query index 1"
        assert failure_details[1] == "Failed to retrieve images: Get failed"


def test_get_studies_from_study_pacs_map_zero_files_retrieved(tmp_path):
    """Test that zero file retrievals are logged and treated as failures."""
    from unittest.mock import patch

    output_dir = str(tmp_path / "output")
    os.makedirs(output_dir, exist_ok=True)

    pacs = PacsConfiguration("localhost", 11112, "ORTHANC")
    study_pacs_map = {
        "1.2.3.4.5": (pacs, 0),
        "1.2.3.4.6": (pacs, 1),
    }

    # Mock get_study to simulate failure with 0 files completed
    with patch('utils.get_study') as mock_get:
        mock_get.side_effect = [
            {"success": False, "num_completed": 0, "num_failed": 0, "num_warning": 0, "message": "Get completed with no sub-operations (no files retrieved)"},
            {"success": True, "num_completed": 5, "num_failed": 0, "num_warning": 0, "message": "Get completed successfully"},
        ]

        successful_gets, failed_query_indices, failure_details = get_studies_from_study_pacs_map(
            study_pacs_map, "TEST_AET", output_dir
        )

        # Zero file retrieval should be treated as failure
        assert successful_gets == 1, "Should have 1 successful get (only the one with files)"
        assert len(failed_query_indices) == 1, "Should have 1 failed query index"
        assert 0 in failed_query_indices, "Query index 0 should have failed (zero files)"
        assert len(failure_details) == 1, "Should have 1 failure detail entry"
        assert 0 in failure_details, "Failure details should contain query index 0"
        assert "no files" in failure_details[0].lower(), "Failure message should mention no files"


def test_get_studies_from_study_pacs_map_exception_handling(tmp_path, caplog):
    """Test that exceptions during get_study don't crash the entire job."""
    from unittest.mock import patch

    output_dir = str(tmp_path / "output")
    os.makedirs(output_dir, exist_ok=True)

    pacs = PacsConfiguration("localhost", 11112, "ORTHANC")
    study_pacs_map = {
        "1.2.3.4.5": (pacs, 0),
        "1.2.3.4.6": (pacs, 1),
        "1.2.3.4.7": (pacs, 2),
    }

    # Mock get_study to simulate various failures including exception
    with patch('utils.get_study') as mock_get:
        mock_get.side_effect = [
            Exception("Network timeout"),  # First call raises exception
            {"success": True, "num_completed": 5, "num_failed": 0, "num_warning": 0, "message": "Get completed"},  # Second succeeds
            {"success": False, "message": "PACS refused connection"},  # Third fails normally
        ]

        successful_gets, failed_query_indices, failure_details = get_studies_from_study_pacs_map(
            study_pacs_map, "TEST_AET", output_dir
        )

        # Job should continue despite exception
        assert successful_gets == 1, "Should have 1 successful get"
        assert len(failed_query_indices) == 2, "Should have 2 failed query indices"
        assert 0 in failed_query_indices, "Query index 0 should have failed (exception)"
        assert 2 in failed_query_indices, "Query index 2 should have failed (normal failure)"
        assert "Exception during retrieval" in failure_details[0], "Should record exception"
        assert "Network timeout" in failure_details[0], "Should include exception message"

        # Check that exception was logged but didn't crash
        assert any("Exception while retrieving" in record.message for record in caplog.records)


def test_find_studies_from_pacs_list_exception_handling(tmp_path, caplog):
    """Test that exceptions during find_studies don't crash the entire job."""
    from unittest.mock import patch

    query_params_list = [
        {"AccessionNumber": "ACC001"},
        {"AccessionNumber": "ACC002"},
        {"AccessionNumber": "ACC003"},
    ]

    # Expected values list is a list of tuples: (expected_accession, query_index)
    expected_values_list = [
        ("ACC001", 0),
        ("ACC002", 1),
        ("ACC003", 2),
    ]

    pacs = PacsConfiguration("localhost", 11112, "ORTHANC")
    pacs_list = [pacs]

    # Mock find_studies to simulate exception on first query, success on second, empty on third
    with patch('utils.find_studies') as mock_find:
        mock_find.side_effect = [
            Exception("Connection refused"),  # First query raises exception
            [{"StudyInstanceUID": "1.2.3.4.5", "AccessionNumber": "ACC002"}],  # Second succeeds
            [],  # Third returns empty
        ]

        study_pacs_map, failed_query_indices, failure_details = find_studies_from_pacs_list(
            pacs_list, query_params_list, "TEST_AET", expected_values_list
        )

        # Job should continue despite exception
        assert len(study_pacs_map) == 1, "Should have 1 study found"
        assert "1.2.3.4.5" in study_pacs_map, "Should have found study from second query"
        assert len(failed_query_indices) == 2, "Should have 2 failed queries"
        assert 0 in failed_query_indices, "Query index 0 should have failed (exception)"
        assert 2 in failed_query_indices, "Query index 2 should have failed (no results)"
        assert "Connection refused" in failure_details[0], "Should include exception message"

        # Check that exception was logged
        assert any("Failed to find studies for query 1" in record.message for record in caplog.records)
